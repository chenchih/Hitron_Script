import openpyxl,string 
from openpyxl.utils import get_column_letter
f=open(r'D:\test\data_out.txt','r+') #open text 
#########if load excel file ########################
#excel=openpyxl.load_workbook(r'D:\\test\\test.xlsx') #open excel 
#excel=openpyxl.load_workbook(r'D:\\test\\test.xlsx') #open excel 
excel = openpyxl.Workbook() 
sheet=excel.worksheets 
line=f.readline(); #read text 
 
while line: 
    list123=[] 
    list123 = line.split(sep=' ') #convert, 
    for i in range(0, len(list123)): # remove space 
        list123[i] = list123[i].strip('\n')
        #print(list[i])
    sheet[0].append(list123) #wrire into excel 
    #sheet[0].column_dimensions.width = 20
    column = 1
    while column < 5:
        i = get_column_letter(column)
        sheet[0].column_dimensions[i].width = 20
        column += 1
  
    line=f.readline() #read next line 
    
excel.save(r'D:\test\test1.xlsx')